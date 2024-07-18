#Abrir un excel vacio
from openpyxl import Workbook
wb = Workbook()
#creacion del objeto Workbook
ws = wb.active
#especificamos
filesheet = "archivos/excel.xlsx"

ws1 = wb.create_sheet("Mysheet")
ws2 = wb.create_sheet("Mysheet",0)
ws3 = wb.create_sheet("Mysheet",-1)

ws.title = "New Title"

ws3 = wb["New Title"]

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)
    
source = wb.active
target = wb.copy_worksheet(source)

ws['A4'] = 4

wb.save(filesheet)