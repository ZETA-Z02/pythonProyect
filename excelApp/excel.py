#EJEMPLO PARA GUARDAR DATOS EN CELDAS EN UN LIBRO EXCEL YA CREADO
# importamos load_workbook
from openpyxl import load_workbook

# ruta de nuestro archivo
filesheet = "archivos/plantilla.xlsx"

# creamos el objeto load_workbook
wb = load_workbook(filesheet)

# Seleccionamos el archivo
sheet = wb.active

celda = 'I15'
# Ingresamos el valor 56 en la celda 'A1'
sheet[celda] = 100

# Ingresamos el valor 1845 en la celda 'B3'
sheet['I13'] = 'correspone'

# Guardamos el archivo con los cambios
wb.save(filesheet)