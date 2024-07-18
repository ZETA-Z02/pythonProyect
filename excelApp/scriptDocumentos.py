#aplicacion para exportar datos de python a excel para facilitar el tipeo de documentos a excel
import locale
#importamos la todo para usar openpyxl
from openpyxl import Workbook, load_workbook

# Configura la localidad para el formato deseado
locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')  # Cambia 'es_ES.UTF-8' a tu localidad deseada
##CREAMOS EL EXCEL, ABRIENDO LA PLANTILLA PARA LLENAR DATOS
# ruta de nuestro archivo
filesheet = "archivos/plantilla.xlsx"
# creamos el objeto load_workbook
wb = load_workbook(filesheet)
# Seleccionamos el archivo
sheet = wb.active

def formatear_numero(numero):
    return locale.format_string("%.2f", numero, grouping=True)

#validacion de datos
#funcion para validar DATOS DE CARACTERES QUE SEAN DEL ALFABETO A-Z==todo a MAYUSCULA
def justString(*datos):
    datos_validados = []
    for dato in datos:
        if dato.isalpha():
            dato.upper()
            datos_validados.append(dato)
            #print("si contiene puras letras: ",dato)     
        else:
            return False
    return datos_validados
#para verificar si son todos numero y que solo tengan 2 flotantes
def justNumber(*numeros,ano='año'):
    if ano =='año':
        numeros_validados = []
        for numero in numeros:
            if numero.isalpha() == False:
                number = round(float(numero), 2)
                numeros_validados.append(number)
                #print("si contiene puras letras: ",dato)     
            else:
                return False
        return numeros_validados
    else:
        try:
            ano_validado = int(ano)
            return ano_validado
        except ValueError:
            return False
            
#---------------------------------------------------------------------------------------
###se pide los primeros datos del titulo
while True:
    #agregar validacion de datos
    nombre = input("Ingrese el nombre: ")
    ano_titulo = input("Ano del titulo del documento: ")
    dependencia = input("Ingrese la dependencia: ")
    condicion = input("La condicion del sujeto: ")
    
    datos_validados = justString(nombre,dependencia,condicion)  
    numeros_validados = justNumber(ano=ano_titulo)
    if datos_validados == False and numeros_validados == False:
        print("fallo, toca volver a poner los datos")
        continue
    else:
        #print(datos_yaValidados)
        break

#print(datos_validados)
#print(numeros_validados)

#---------------------------------------------------------------------------------------------
while True:
    try:
        tiempoDeServicio = int(input("Cuantos años trabajo: "))
        ano_inicio = int(input("Desde que año inicio: "))
        break
    except ValueError:
        print("Error: Ingresa un número válido.")
#-------------------------------------------------------------------------------------------------

#todos los meses que va recorer para llenar datos
meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Setiembre','Octubre','Noviembre','Diciembre',]
datosLlenar = ['CARGO','CONDICION','GRADO SUB GRADO','BASICA IMPORTE','TOTAL REMUNERACION','D.L. 19990','FONAVI']
data = []
n = 0
#pasar por cada año y llenar los datos correspondientes
while n<tiempoDeServicio:
    print('Año del que se esta llenando los datos: ',ano_inicio)
    for mes in meses:
        print("Ingrese todos los datos correspondientes del siguietne mes: ",mes)
        #H-I-J-K-L-M-N
        fila = 13
        columnas = ['H','I','J','K','L','M','N']
        while True:
            zipVar = zip(datosLlenar,columnas)
            print(type(zipVar))
            continuar = input('yes or not?: ')
            if continuar == 'n':
                break
        # for llenar, columna in zip(datosLlenar, columnas):
        #     print("llenara los datos de: ", llenar)
        #     datos = input("Ingrese los datos correspondiendes: ")
        #     columna+=str(fila)
        #     sheet[columna] = datos  
        #     fila+=1
    ano_inicio+=1
    n+=1
    
##se descomentara esto cuando los datos se puedan enviar correctamente para crear el excel
# Guardamos el archivo con los cambios
wb.save(filesheet)









